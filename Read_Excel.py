import openpyxl 
import unittest

wb = openpyxl.load_workbook('ImportTemplate.xlsx', data_only=True)

ws = wb.active

#search = int(input("select designation ID: "))

ws = wb["April 2nd"]

def find_an_email(search):
    count = 0
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=12, max_col=3): 
        count = 0
        for cell in row:
            if cell.value == search:
                            return str(ws["I"+str(count)].value)
            else:
                continue
    
#print(find_an_email(search))

def send_email(email):
      print(email)

def send_emails():
      count = 1
      for row in ws.iter_rows():
            count += 1
            if ws["I"+str(count)].value == None:
                  return 
            else:
                  email = str(ws["I"+str(count)].value)
                  send_email(email)

send_emails()
    
